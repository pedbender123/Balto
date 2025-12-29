import os
import requests
import openpyxl
import base64
import json
import time
import wave

# Configurações
# Default to Localhost for local testing ease, per user request. 
# Can be overridden by BALTO_SERVER_URL env var.
SERVER_URL = os.environ.get("BALTO_SERVER_URL", "http://localhost:8765") 
print(f"Using Server URL: {SERVER_URL}")

INPUT_DIR = "testes/1_input"
OUTPUT_DIR = "testes/planilhas"
SEGMENTS_TEMP = "testes/2_cortados"

# Arquivos de Saída
FILE_ORIGINAIS = os.path.join(OUTPUT_DIR, "Relatorio_Originais.xlsx")
FILE_SEGMENTOS = os.path.join(OUTPUT_DIR, "Relatorio_Segmentos.xlsx")

HEADERS_ORIGINAIS = [
    'Nome do Arquivo',
    'Transcrição Completa (ElevenLabs)'
]

HEADERS_SEGMENTOS = [
    'Arquivo Original', 
    'Nome do Segmento', 
    'Duração (s)', 
    'Transcrição ElevenLabs (Segmento)', 
    'Transcrição AssemblyAI (Segmento)',
    'Transcrição Deepgram (Segmento)',
    'Transcrição Gladia (Segmento)', 
    'Qualidade', 
    'Acurácia'
]

def ensure_dirs():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(SEGMENTS_TEMP, exist_ok=True)

def load_or_create_wb(filepath, headers):
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        return wb
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        return wb

def get_existing_transcriptions(wb):
    """Carrega transcrições já feitas para memória (cache)."""
    cache = {}
    ws = wb.active
    # Assume row 1 = headers
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]: # Filename
            transcript = row[1]
            # Ignore errors in cache so we retry them
            if transcript and not str(transcript).startswith("[ERROR") and not str(transcript).startswith("[EXCEPTION"):
                cache[row[0]] = transcript
    return cache

def transcribe_file(filepath, provider):
    try:
        with open(filepath, 'rb') as f:
            files = {
                'audio': (os.path.basename(filepath), f, 'audio/wav'), # Server handles ext check/decoding
                'provider': (None, provider)
            }
            res = requests.post(f"{SERVER_URL}/api/test/transcrever", files=files, timeout=120)
            if res.status_code == 200:
                return res.json().get("texto", "")
            else:
                return f"[ERROR {res.status_code}] {res.text}"
    except Exception as e:
        return f"[EXCEPTION] {str(e)}"

def segment_file(filepath):
    try:
        with open(filepath, 'rb') as f:
            files = {'audio': (os.path.basename(filepath), f, 'audio/webm')}
            res = requests.post(f"{SERVER_URL}/api/test/segmentar", files=files, timeout=120)
            if res.status_code == 200:
                return res.json().get("segments", [])
            else:
                print(f"Error segmenting: {res.text}")
                return []
    except Exception as e:
        print(f"Exception segmenting: {e}")
        return []

def main():
    ensure_dirs()
    
    # 1. Carregar/Criar Planilha de Originais (Cache)
    wb_orig = load_or_create_wb(FILE_ORIGINAIS, HEADERS_ORIGINAIS)
    cache_orig = get_existing_transcriptions(wb_orig)
    ws_orig = wb_orig.active

    # 2. Carregar/Criar Planilha de Segmentos (Modo Resume)
    wb_seg = load_or_create_wb(FILE_SEGMENTOS, HEADERS_SEGMENTOS)
    ws_seg = wb_seg.active
    
    # Identificar quais arquivos JÁ tem segmentos na planilha para pular
    existing_files_in_seg = set()
    for row in ws_seg.iter_rows(min_row=2, values_only=True):
        if row[0]: # Coluna 'Arquivo Original'
            existing_files_in_seg.add(row[0])
            
    print(f"Arquivos já processados (Segmentos): {len(existing_files_in_seg)}")
    
    input_files = sorted([f for f in os.listdir(INPUT_DIR) if not f.startswith('.')])
    print(f"Arquivos encontrados: {len(input_files)}")

    for idx, filename in enumerate(input_files):
        print(f"\n[{idx+1}/{len(input_files)}] Processando: {filename}")
        
        if filename in existing_files_in_seg:
            print(f" -> PULANDO: Já existe no relatório de segmentos.")
            continue

        original_path = os.path.join(INPUT_DIR, filename)
        
        # --- ETAPA 1: Transcrição do Original (Com Cache) ---
        full_transcript = cache_orig.get(filename)
        
        if full_transcript:
            print(f" -> Original já transcrito (Cache).")
        else:
            print(f" -> Transcrevendo Original (ElevenLabs)...")
            full_transcript = transcribe_file(original_path, "elevenlabs")
            
            # Salvar no cache e na planilha imediatamente
            ws_orig.append([filename, full_transcript])
            wb_orig.save(FILE_ORIGINAIS) # Save checkpoint
            cache_orig[filename] = full_transcript
            print(f" -> Transcrição salva.")

        # --- ETAPA 2: Segmentação ---
        print(" -> Segmentando...")
        segments = segment_file(original_path)
        print(f"    {len(segments)} segmentos encontrados.")
        
        for i, seg in enumerate(segments):
            seg_name = f"{os.path.splitext(filename)[0]}_seg_{i:03d}.wav"
            seg_path = os.path.join(SEGMENTS_TEMP, seg_name)
            
            raw_bytes = base64.b64decode(seg["audio_base64"])
            
            # Salvar como WAV valido
            with wave.open(seg_path, 'wb') as wf:
                wf.setnchannels(1)
                wf.setsampwidth(2)
                wf.setframerate(16000)
                wf.writeframes(raw_bytes)
            
            duration = seg.get("duration_sec", 0)
            
            print(f"    -> Seg {i}: Transcrevendo (Eleven, Assembly, Deepgram, Gladia)...")
            text_eleven = transcribe_file(seg_path, "elevenlabs")
            text_assembly = transcribe_file(seg_path, "assemblyai")
            text_deepgram = transcribe_file(seg_path, "deepgram")
            text_gladia = transcribe_file(seg_path, "gladia")
            
            row = [
                filename,
                seg_name,
                f"{duration:.2f}",
                text_eleven,
                text_assembly,
                text_deepgram,
                text_gladia,
                "", # Qualidade
                ""  # Acuracia
            ]
            ws_seg.append(row)
        
        # Salvar planilha de segmentos a cada arquivo processado para não perder progresso
        wb_seg.save(FILE_SEGMENTOS)
        
        print("\n[PILOT TEST] Stopping after 1 file.")
        break

    print(f"\nConcluído!")
    print(f"Relatório Originais: {FILE_ORIGINAIS}")
    print(f"Relatório Segmentos: {FILE_SEGMENTOS}")

if __name__ == "__main__":
    main()
