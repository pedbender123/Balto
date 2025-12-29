import os
import requests
import openpyxl
from dotenv import load_dotenv

# Carregar variáveis de ambiente (chave nova local)
load_dotenv(os.path.join(os.path.dirname(__file__), "../backend/.env"))
API_KEY = os.environ.get("ELEVENLABS_API_KEY")

if not API_KEY:
    print("ERRO: ELEVENLABS_API_KEY não encontrada no ../backend/.env")
    exit(1)

INPUT_DIR = "testes/1_input"
OUTPUT_FILE = "testes/planilhas/Relatorio_Originais.xlsx"

def transcribe_elevenlabs(filepath):
    url = "https://api.elevenlabs.io/v1/speech-to-text"
    headers = {
        "xi-api-key": API_KEY
    }
    
    # Modelo Scribe en-v1 ou similar? O endpoint padrão da ElevenLabs STT (Scribe)
    # Documentação atual: POST /v1/speech-to-text
    # Body: file/audio, model_id (optional)
    
    try:
        with open(filepath, "rb") as f:
            files = {"file": f}
            data = {"model_id": "scribe_v1"} # Exemplo, ou default
            
            response = requests.post(url, headers=headers, files=files, data=data, timeout=300)
            
        if response.status_code == 200:
            return response.json().get("text", "")
        else:
            return f"[ERROR {response.status_code}] {response.text}"
    except Exception as e:
        return f"[EXCEPTION] {str(e)}"

def main():
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    
    # Carregar ou Criar Planilha
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        existing_files = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nome do Arquivo", "Transcrição Completa (ElevenLabs)"])
        existing_files = set()

    files = sorted([f for f in os.listdir(INPUT_DIR) if not f.startswith('.')])
    
    print(f"Encontrados {len(files)} arquivos.")
    print(f"Chave API: {API_KEY[:5]}...{API_KEY[-5:]}")
    
    for idx, filename in enumerate(files):
        print(f"[{idx+1}/{len(files)}] {filename}", end=" ", flush=True)
        
        if filename in existing_files:
            print("-> JÁ PROCESSADO.")
            continue
            
        filepath = os.path.join(INPUT_DIR, filename)
        
        print("-> Transcrevendo...", end=" ", flush=True)
        transcript = transcribe_elevenlabs(filepath)
        
        ws.append([filename, transcript])
        wb.save(OUTPUT_FILE)
        
        status = "SUCESSO" if not transcript.startswith("[") else "FALHA"
        print(f"-> {status}")

    print("\nProcesso concluído! Planilha salva.")

if __name__ == "__main__":
    main()
